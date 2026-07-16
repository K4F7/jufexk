from __future__ import annotations

import argparse
import csv
import json
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from pipeline import normalize


VALID_DECISIONS = {"", "approve", "reject", "skip"}


def read_alias_decisions(workbook_path: Path) -> list[dict[str, str]]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        if "OCR课程别名核对" not in workbook.sheetnames:
            raise ValueError("workbook does not contain OCR课程别名核对")
        sheet = workbook["OCR课程别名核对"]
        headers = [str(cell.value or "").strip() for cell in sheet[2]]
        required = {"ocr_course_name", "candidate_code", "candidate_name", "decision"}
        if not required.issubset(headers):
            raise ValueError(f"alias sheet is missing columns: {sorted(required - set(headers))}")
        rows: list[dict[str, str]] = []
        for values in sheet.iter_rows(min_row=3, values_only=True):
            row = {headers[index]: str(value or "").strip() for index, value in enumerate(values) if index < len(headers)}
            if not row.get("ocr_course_name"):
                continue
            decision = row.get("decision", "").lower()
            if decision not in VALID_DECISIONS:
                raise ValueError(f"invalid decision for {row['ocr_course_name']}: {decision}")
            row["decision"] = decision
            rows.append(row)
        return rows
    finally:
        workbook.close()


def approved_aliases(
    decisions: list[dict[str, str]], reference: dict[str, Any]
) -> dict[str, dict[str, Any]]:
    catalog = {
        (str(row.get("code", "")), str(row.get("name", ""))): row
        for row in reference.get("courses", [])
    }
    approved: dict[str, dict[str, Any]] = {}
    for row in decisions:
        if row["decision"] != "approve":
            continue
        source_key = normalize(row["ocr_course_name"])
        target_key = (row["candidate_code"], row["candidate_name"])
        if target_key not in catalog:
            raise ValueError(
                f"approved alias target is not in catalog: {row['ocr_course_name']} -> {target_key}"
            )
        if source_key in approved:
            raise ValueError(f"multiple approved targets for OCR course: {row['ocr_course_name']}")
        approved[source_key] = catalog[target_key]
    return approved


def apply_aliases(
    preview_rows: list[dict[str, str]],
    aliases: dict[str, dict[str, Any]],
    reference: dict[str, Any],
) -> tuple[list[dict[str, str]], int]:
    linked = {
        (int(row["course_id"]), int(row["teacher_id"]))
        for row in reference.get("course_teachers", [])
    }
    changed = 0
    output: list[dict[str, str]] = []
    for source in preview_rows:
        row = source.copy()
        course = aliases.get(normalize(row.get("ocr_course_name", "")))
        if not course:
            output.append(row)
            continue
        course_id = int(course["id"])
        row["matched_course_id"] = str(course_id)
        row["matched_course_name"] = str(course["name"])
        row["course_match_score"] = "1.0"
        reasons = [
            reason
            for reason in row.get("review_reason", "").split(";")
            if reason and reason not in {"course_unmatched_or_ambiguous", "teacher_not_linked_to_course"}
        ]
        teacher_id = int(row["matched_teacher_id"]) if row.get("matched_teacher_id") else 0
        if teacher_id and (course_id, teacher_id) not in linked:
            reasons.append("teacher_not_linked_to_course")
        reasons.append("course_alias_human_confirmed")
        row["review_reason"] = ";".join(dict.fromkeys(reasons))
        row["needs_review"] = "True"
        changed += 1
        output.append(row)
    return output, changed


def read_csv(path: Path) -> tuple[list[str], list[dict[str, str]]]:
    with path.open(encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        return list(reader.fieldnames or []), list(reader)


def write_csv(path: Path, fields: list[str], rows: list[dict[str, str]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def main() -> int:
    parser = argparse.ArgumentParser(description="Apply explicitly approved course aliases to an OCR preview")
    parser.add_argument("--workbook", required=True)
    parser.add_argument("--preview", required=True)
    parser.add_argument("--reference", required=True)
    parser.add_argument("--out", required=True)
    parser.add_argument("--report", required=True)
    args = parser.parse_args()

    decisions = read_alias_decisions(Path(args.workbook))
    reference = json.loads(Path(args.reference).read_text(encoding="utf-8-sig"))
    aliases = approved_aliases(decisions, reference)
    fields, preview = read_csv(Path(args.preview))
    output, changed = apply_aliases(preview, aliases, reference)
    write_csv(Path(args.out), fields, output)
    report = {
        "approved_alias_count": len(aliases),
        "updated_review_count": changed,
        "input_review_count": len(preview),
        "output": str(Path(args.out)),
    }
    report_path = Path(args.report)
    report_path.parent.mkdir(parents=True, exist_ok=True)
    report_path.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(report, ensure_ascii=False))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
